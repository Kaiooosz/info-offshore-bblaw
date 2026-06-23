#!/usr/bin/env python3
"""Gera a planilha Capital Blindado — Calculadora de Estruturas."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "entregaveis", "Capital-Blindado-Calculadora.xlsx")

BG_DARK = "070D0A"
BG_CARD = "0E1A14"
BG_INPUT = "1A2E1F"
WHITE = "FFFFFF"
GRAY = "999999"
GREEN_DIM = "2D5A3D"

font_title = Font(name="Arial", size=16, bold=True, color=WHITE)
font_h2 = Font(name="Arial", size=12, bold=True, color=WHITE)
font_h3 = Font(name="Arial", size=10, bold=True, color="CCCCCC")
font_label = Font(name="Arial", size=10, color="CCCCCC")
font_input = Font(name="Arial", size=11, bold=True, color=WHITE)
font_result = Font(name="Arial", size=12, bold=True, color=WHITE)
font_result_lg = Font(name="Arial", size=14, bold=True, color=WHITE)
font_small = Font(name="Arial", size=9, color=GRAY)
font_header = Font(name="Arial", size=9, bold=True, color=WHITE)
font_cell = Font(name="Arial", size=9, color="CCCCCC")

fill_dark = PatternFill(start_color=BG_DARK, end_color=BG_DARK, fill_type="solid")
fill_card = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type="solid")
fill_input = PatternFill(start_color=BG_INPUT, end_color=BG_INPUT, fill_type="solid")
fill_header = PatternFill(start_color=GREEN_DIM, end_color=GREEN_DIM, fill_type="solid")

align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="1A2E1F"),
    right=Side(style="thin", color="1A2E1F"),
    top=Side(style="thin", color="1A2E1F"),
    bottom=Side(style="thin", color="1A2E1F"),
)

BRL = '#,##0'
BRL_DEC = '#,##0.00'
PCT = '0.0%'
MESES = '0 "meses"'


def apply_bg(ws, rows, cols):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = fill_dark


def build():
    wb = Workbook()

    # ── ABA 1: SIMULADOR ──
    ws = wb.active
    ws.title = "Simulador"
    ws.sheet_properties.tabColor = "2D5A3D"

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 22

    apply_bg(ws, 55, 9)

    # Titulo
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "CAPITAL BLINDADO — SIMULADOR"
    c.font = font_title
    c.alignment = align_left

    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = "Bezerra Borges Advogados / bezerraborges.com.br"
    c.font = font_small
    c.alignment = align_left

    # SECAO: SEUS DADOS
    ws["B5"].value = "SEUS DADOS"
    ws["B5"].font = font_h2

    labels_inputs = [
        (7, "Patrimonio total (R$)", 3000000),
        (8, "Renda anual (R$)", 600000),
        (9, "Impostos pagos por ano (R$)", 180000),
        (10, "Custo estimado da estrutura (R$)", 25000),
    ]
    for row, label, default in labels_inputs:
        ws.cell(row=row, column=2, value=label).font = font_label
        cell = ws.cell(row=row, column=3, value=default)
        cell.font = font_input
        cell.fill = fill_input
        cell.number_format = BRL
        cell.border = thin_border
        cell.alignment = align_right

    # SECAO: RESULTADOS
    ws["B12"].value = "RESULTADOS"
    ws["B12"].font = font_h2

    results = [
        (14, "Economia tributaria estimada (ano)", "=C9*0.35", BRL),
        (15, "Economia em 5 anos", "=C14*5", BRL),
        (16, "Economia em 10 anos", "=C14*10", BRL),
        (17, "Payback (meses)", '=IF(C14>0,ROUNDUP(C10/C14*12,0),"N/A")', MESES),
        (18, "ROI no primeiro ano", "=IF(C10>0,(C14-C10)/C10,0)", PCT),
        (19, "ROI em 5 anos", "=IF(C10>0,(C15-C10)/C10,0)", PCT),
    ]
    for row, label, formula, fmt in results:
        ws.cell(row=row, column=2, value=label).font = font_label
        cell = ws.cell(row=row, column=3)
        cell.value = formula
        cell.font = font_result
        cell.fill = fill_card
        cell.number_format = fmt
        cell.border = thin_border
        cell.alignment = align_right

    # SECAO: COMPARATIVO POR ESTRUTURA
    ws["E5"].value = "COMPARATIVO"
    ws["E5"].font = font_h2

    headers = ["Estrutura", "Setup", "Manut./ano", "Economia est.", "Payback"]
    header_cols = [2, 3, 4, 5, 6]
    row_start = 22

    ws["B21"].value = "COMPARATIVO POR ESTRUTURA"
    ws["B21"].font = font_h2

    for i, (col, header) in enumerate(zip(header_cols, headers)):
        cell = ws.cell(row=row_start, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    estruturas = [
        ("LLC EUA (Wyoming)", 3500, 1500, "=C9*0.15"),
        ("LLC EUA (Delaware)", 5500, 2000, "=C9*0.15"),
        ("Holding Portugal", 35000, 15000, "=C9*0.25"),
        ("Holding Malta", 22000, 12000, "=C9*0.30"),
        ("Nevis LLC", 12000, 6000, "=C9*0.20"),
        ("BVI Business Co.", 10000, 5000, "=C9*0.20"),
        ("Cayman Exempted", 35000, 18000, "=C9*0.35"),
        ("Trust Jersey", 80000, 25000, "=C9*0.35"),
        ("Trust Cayman", 70000, 22000, "=C9*0.35"),
        ("Fundacao Liechtenstein", 100000, 30000, "=C9*0.35"),
        ("Foundation Panama", 25000, 8000, "=C9*0.25"),
    ]

    for i, (nome, setup, manut, econ_formula) in enumerate(estruturas):
        r = row_start + 1 + i
        ws.cell(row=r, column=2, value=nome).font = font_cell
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=2).fill = fill_card

        c_setup = ws.cell(row=r, column=3, value=setup)
        c_setup.font = font_cell
        c_setup.number_format = BRL
        c_setup.border = thin_border
        c_setup.fill = fill_card
        c_setup.alignment = align_right

        c_manut = ws.cell(row=r, column=4, value=manut)
        c_manut.font = font_cell
        c_manut.number_format = BRL
        c_manut.border = thin_border
        c_manut.fill = fill_card
        c_manut.alignment = align_right

        c_econ = ws.cell(row=r, column=5)
        c_econ.value = econ_formula
        c_econ.font = font_cell
        c_econ.number_format = BRL
        c_econ.border = thin_border
        c_econ.fill = fill_card
        c_econ.alignment = align_right

        setup_ref = get_column_letter(3) + str(r)
        econ_ref = get_column_letter(5) + str(r)
        c_pay = ws.cell(row=r, column=6)
        c_pay.value = f'=IF({econ_ref}>0,ROUNDUP({setup_ref}/{econ_ref}*12,0),"N/A")'
        c_pay.font = font_cell
        c_pay.number_format = MESES
        c_pay.border = thin_border
        c_pay.fill = fill_card
        c_pay.alignment = align_right

    # Nota
    note_row = row_start + len(estruturas) + 2
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)
    c = ws.cell(row=note_row, column=2)
    c.value = "Valores estimados para fins educacionais. Consulte um profissional para analise do seu caso."
    c.font = font_small
    c.alignment = align_left

    # SECAO: PROJECAO 5 E 10 ANOS
    proj_row = note_row + 3
    ws.cell(row=proj_row, column=2, value="PROJECAO DE ECONOMIA").font = font_h2

    proj_headers = ["Ano", "Economia acumulada", "Custo acumulado", "Resultado liquido"]
    for i, h in enumerate(proj_headers):
        cell = ws.cell(row=proj_row + 1, column=2 + i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for yr in range(1, 11):
        r = proj_row + 1 + yr
        ws.cell(row=r, column=2, value=yr).font = font_cell
        ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=2).fill = fill_card
        ws.cell(row=r, column=2).alignment = align_center

        c_econ = ws.cell(row=r, column=3)
        c_econ.value = f"=C14*{yr}"
        c_econ.font = font_cell
        c_econ.number_format = BRL
        c_econ.border = thin_border
        c_econ.fill = fill_card
        c_econ.alignment = align_right

        c_custo = ws.cell(row=r, column=4)
        if yr == 1:
            c_custo.value = "=C10"
        else:
            c_custo.value = f"=C10+({yr}-1)*C10*0.1"
        c_custo.font = font_cell
        c_custo.number_format = BRL
        c_custo.border = thin_border
        c_custo.fill = fill_card
        c_custo.alignment = align_right

        c_liq = ws.cell(row=r, column=5)
        c_liq.value = f"={get_column_letter(3)}{r}-{get_column_letter(4)}{r}"
        c_liq.font = font_cell
        c_liq.number_format = BRL
        c_liq.border = thin_border
        c_liq.fill = fill_card
        c_liq.alignment = align_right

    # ── ABA 2: INSTRUCOES ──
    ws2 = wb.create_sheet("Instrucoes")
    ws2.sheet_properties.tabColor = "555555"
    ws2.column_dimensions["B"].width = 80
    apply_bg(ws2, 25, 5)

    ws2["B2"].value = "COMO USAR ESTA PLANILHA"
    ws2["B2"].font = font_title

    instrucoes = [
        "1. Acesse a aba 'Simulador'",
        "2. Preencha os 4 campos em verde com os seus dados reais:",
        "   - Patrimonio total: some imoveis, investimentos, participacoes e outros ativos",
        "   - Renda anual: inclua pro-labore, dividendos, alugueis e outras fontes",
        "   - Impostos pagos: o total de IR, CSLL, ITCMD e outros tributos que voce paga por ano",
        "   - Custo da estrutura: o investimento estimado para montar a estrutura (setup)",
        "3. Os resultados sao calculados automaticamente:",
        "   - Economia tributaria estimada (35% dos impostos atuais como referencia)",
        "   - Payback: quantos meses para a estrutura se pagar",
        "   - ROI: retorno sobre o investimento no primeiro ano e em 5 anos",
        "4. O comparativo mostra custos e payback para cada tipo de estrutura",
        "5. A projecao mostra economia acumulada em 10 anos",
        "",
        "IMPORTANTE: Os calculos sao estimativas educacionais.",
        "A economia real depende do seu caso especifico.",
        "Consulte a Bezerra Borges Advogados para uma analise personalizada.",
        "",
        "bezerraborges.com.br",
    ]
    for i, txt in enumerate(instrucoes):
        cell = ws2.cell(row=4 + i, column=2, value=txt)
        cell.font = font_label if not txt.startswith("IMPORTANTE") else font_h3
        cell.alignment = align_wrap

    wb.save(OUT)
    print(f"Planilha gerada: {OUT}")


if __name__ == "__main__":
    build()
